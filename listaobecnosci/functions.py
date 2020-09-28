from datetime import datetime
from math import ceil
from openpyxl import load_workbook
from reportlab.lib import colors
from template import dataToTemplate
from tableWorkersInPdf import genPinTable
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate

LIBRE_OFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"


class AttendanceList:

    @staticmethod
    def loadDate(file_workers_list, zmiana, month):
        global zmiana_m, zmiana_w
        workbook = load_workbook(filename=file_workers_list)
        sheet = workbook.active
        date_workers = []
        women = []
        men = []
        for a in range(1, 100):
            dana = str(sheet["A" + str(a)].value)
            typ = str(sheet['B' + str(a)].value)
            if dana == 'None':
                break
            else:
                date_workers.append((dana, typ))

        for worker, typ in date_workers:
            if typ == 'k':
                women.append(worker)
            elif typ == 'm':
                men.append(worker)
            else:
                continue

        womenList = women
        menList = men

        if zmiana == '1':
            zmiana_w = 'I'
            zmiana_m = 'II'

            for t in range(1, ceil(len(women) / 3) + 1):
                print(ceil(len(women) / 3))

                try:
                    dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                except IndexError:
                    if len(women) == 1:
                        women.append("")
                        women.append("")
                        dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                        women.clear()

                        continue
                    else:
                        women.append("")
                        dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                        women.clear()
                        continue

                for _ in range(3):
                    del women[0]

            for t in range(1, ceil(len(men) / 3) + 1):
                print(ceil(len(men) / 3))

                try:
                    dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                except IndexError:
                    if len(men) == 1:
                        men.append("")
                        men.append("")
                        dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                        men.clear()

                        continue
                    else:
                        men.append("")
                        dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                        men.clear()
                        continue

                for _ in range(3):
                    del men[0]

        elif zmiana == '2':

            zmiana_w = 'III'
            zmiana_m = 'IV'

            for t in range(1, ceil(len(women) / 3) + 1):
                print(ceil(len(women) / 3))

                try:
                    dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                except IndexError:
                    if len(women) == 1:
                        women.append("")
                        women.append("")
                        dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                        women.clear()

                        continue
                    else:
                        women.append("")
                        dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                        women.clear()
                        continue

                for _ in range(3):
                    del women[0]

            for t in range(1, ceil(len(men) / 3) + 1):
                print(ceil(len(men) / 3))

                try:
                    dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                except IndexError:
                    if len(men) == 1:
                        men.append("")
                        men.append("")
                        dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                        men.clear()

                        continue
                    else:
                        men.append("")
                        dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                        men.clear()
                        continue

                for _ in range(3):
                    del men[0]
        else:
            print("błędnie podano nr zmiany")

        return womenList, menList, zmiana_w, zmiana_m

    @staticmethod
    def loadDateShort(file_workers_list, zmiana, month):
        global zmiana_m, zmiana_w
        workbook = load_workbook(filename=file_workers_list)
        sheet = workbook.active
        date_workers = []
        women = []
        men = []
        for a in range(1, 100):
            dana = str(sheet["A" + str(a)].value)
            typ = str(sheet['B' + str(a)].value)
            if dana == 'None':
                break
            else:
                date_workers.append((dana, typ))

        for worker, typ in date_workers:
            if typ == 'k':
                women.append(worker)
            elif typ == 'm':
                men.append(worker)
            else:
                continue

        womenList = women
        menList = men
        if zmiana == '1':
            zmiana_w = 'I'
            zmiana_m = 'II'
        elif zmiana == '2':

            zmiana_w = 'III'
            zmiana_m = 'IV'

        return womenList, menList, zmiana_w, zmiana_m

    @staticmethod
    def markDay(lenMonth, month):
        mark_days_list = []
        ts = []
        for i in range(1, lenMonth + 1):
            x = datetime(2020, month, i)
            day = x.strftime("%A")
            if 'Saturday' == day or 'Sunday' == day:
                ts = ('BACKGROUND', (0, i - 1), (-1, i - 1), colors.darkgrey)
                mark_days_list.append(ts)

        return mark_days_list



    @staticmethod
    def generatorList(listPerson, change, otherParm):

        for t in range(1, ceil(len(listPerson) / 3) + 1):

            try:

                List_person = genPinTable(listPerson[0], listPerson[1], listPerson[2], change, otherParm[1], t,
                                      otherParm[2],
                                      otherParm[3])
                fileName = 'listaobecnosci' + str(otherParm[0]) + str(t) + '.pdf'
                pdf = SimpleDocTemplate(fileName, pagesize=letter, bottomMargin=0, topMargin=5, rightMargin=0,
                                        leftMargin=0)

                elems = []
                elems.append(List_person)
                pdf.build(elems)
            except IndexError:
                if len(listPerson) == 1:
                    listPerson.append("")
                    listPerson.append("")
                    List_person = genPinTable(listPerson[0], listPerson[1], listPerson[2], change, otherParm[1], t,
                                              otherParm[2],
                                              otherParm[3])
                    fileName = 'listaobecnosci' + str(otherParm[0]) + str(t) + '.pdf'
                    pdf = SimpleDocTemplate(fileName, pagesize=letter, bottomMargin=0, topMargin=5, rightMargin=0,
                                            leftMargin=0)

                    elems = []
                    elems.append(List_person)
                    pdf.build(elems)
                    listPerson.clear()

                    continue
                else:
                    listPerson.append("")
                    List_person = genPinTable(listPerson[0], listPerson[1], listPerson[2], change, otherParm[1], t,
                                              otherParm[2],
                                              otherParm[3])
                    fileName = 'listaobecnosci' + str(otherParm[0]) + str(t) + '.pdf'
                    pdf = SimpleDocTemplate(fileName, pagesize=letter, bottomMargin=0, topMargin=5, rightMargin=0,
                                            leftMargin=0)

                    elems = []
                    elems.append(List_person)
                    pdf.build(elems)
                    listPerson.clear()
                    continue

            for _ in range(3):
                del listPerson[0]
        return 0


