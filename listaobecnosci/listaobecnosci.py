from math import ceil

from openpyxl import load_workbook

from template import dataToTemplate
from listaobecnoscipdf import genPinTable

LIBRE_OFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"


class AttendanceList:

    @staticmethod
    def loadDate(file_workers_list, zmiana, month):
        global zmiana_m, zmiana_w
        workbook = load_workbook(filename=file_workers_list)
        sheet = workbook.active
        date_workers = []
        women = []
        men = []
        for a in range(1, 100):
            dana = str(sheet["A" + str(a)].value)
            typ = str(sheet['B' + str(a)].value)
            if dana == 'None':
                break
            else:
                date_workers.append((dana, typ))

        for worker, typ in date_workers:
            if typ == 'k':
                women.append(worker)
            elif typ == 'm':
                men.append(worker)
            else:
                continue

        womenList = women
        menList = men

        if zmiana == '1':
            zmiana_w = 'I'
            zmiana_m = 'II'

            for t in range(1, ceil(len(women) / 3) + 1):
                print(ceil(len(women) / 3))

                try:
                    dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                except IndexError:
                    if len(women) == 1:
                        women.append("")
                        women.append("")
                        dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                        women.clear()

                        continue
                    else:
                        women.append("")
                        dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                        women.clear()
                        continue

                for _ in range(3):
                    del women[0]

            for t in range(1, ceil(len(men) / 3) + 1):
                print(ceil(len(men) / 3))

                try:
                    dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                except IndexError:
                    if len(men) == 1:
                        men.append("")
                        men.append("")
                        dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                        men.clear()

                        continue
                    else:
                        men.append("")
                        dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                        men.clear()
                        continue

                for _ in range(3):
                    del men[0]

        elif zmiana == '2':

            zmiana_w = 'III'
            zmiana_m = 'IV'

            for t in range(1, ceil(len(women) / 3) + 1):
                print(ceil(len(women) / 3))

                try:
                    dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                except IndexError:
                    if len(women) == 1:
                        women.append("")
                        women.append("")
                        dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                        women.clear()

                        continue
                    else:
                        women.append("")
                        dataToTemplate(zmiana_w, "szablon_listy_obecnosci.docx", t, month, women)
                        women.clear()
                        continue

                for _ in range(3):
                    del women[0]

            for t in range(1, ceil(len(men) / 3) + 1):
                print(ceil(len(men) / 3))

                try:
                    dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                except IndexError:
                    if len(men) == 1:
                        men.append("")
                        men.append("")
                        dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                        men.clear()

                        continue
                    else:
                        men.append("")
                        dataToTemplate(zmiana_m, "szablon_listy_obecnosci.docx", t, month, men)
                        men.clear()
                        continue

                for _ in range(3):
                    del men[0]
        else:
            print("błędnie podano nr zmiany")

        return womenList, menList,  zmiana_w, zmiana_m

    @staticmethod
    def loadDateShort(file_workers_list, zmiana, month):
        global zmiana_m, zmiana_w
        workbook = load_workbook(filename=file_workers_list)
        sheet = workbook.active
        date_workers = []
        women = []
        men = []
        for a in range(1, 100):
            dana = str(sheet["A" + str(a)].value)
            typ = str(sheet['B' + str(a)].value)
            if dana == 'None':
                break
            else:
                date_workers.append((dana, typ))

        for worker, typ in date_workers:
            if typ == 'k':
                women.append(worker)
            elif typ == 'm':
                men.append(worker)
            else:
                continue

        womenList = women
        menList = men
        if zmiana == '1':
            zmiana_w = 'I'
            zmiana_m = 'II'
        elif zmiana == '2':

            zmiana_w = 'III'
            zmiana_m = 'IV'

        return womenList, menList, zmiana_w, zmiana_m
